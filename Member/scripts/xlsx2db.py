import os
import subprocess
import logging
import json
import argparse
from openpyxl import load_workbook

def _load_config():
    cfg_file = os.path.join(os.path.dirname(__file__), 'config.json')
    if os.path.exists(cfg_file):
        try:
            with open(cfg_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

cfg_dict = _load_config()
# 优先配置logpath，没有就默认路径
if cfg_dict.get("logpath"):
    log_path = cfg_dict["logpath"]
else:
    log_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "data", "run.log"))

def excel_to_db(xlsx_path: str, db_path: str, table_name: str):
    """
    读取Excel逐行：查→不存在新增→统一修改
    :param xlsx_path: Excel文件路径
    :param db_path: sqlite路径(本函数不用，子脚本自行读取配置)
    :param table_name: 预留表名
    """
    if not os.path.exists(xlsx_path):
        logging.error(f"xlsx文件不存在：{xlsx_path}")
        return

    wb = load_workbook(xlsx_path)
    ws = wb.active
    # 读取首行表头
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # 建立字段下标映射
    col_map = {name: idx for idx, name in enumerate(header)}
    need_cols = ["gaijin_id", "name", "state", "join_date", "landforce", "airforce", "navy"]

    # 字段缺失校验
    for col in need_cols:
        if col not in col_map:
            logging.error(f"Excel缺少必填列：{col}")
            return

    # 逐行循环(从第2行数据开始)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        gaijin_id = row[col_map["gaijin_id"]]
        name      = row[col_map["name"]]
        state     = row[col_map["state"]]
        join_date = row[col_map["join_date"]]
        landforce = row[col_map["landforce"]]
        airforce  = row[col_map["airforce"]]
        navy      = row[col_map["navy"]]

        # 主键为空跳过本行
        if not gaijin_id:
            logging.info(f"行={row_idx} id=- 操作=跳过 原因=盖金号为空")
            continue

        # 为每行汇总状态，最终只输出一行结构化日志
        row_status = {
            'row': row_idx,
            'id': gaijin_id,
            'created': False,
            'modified': False,
            'skipped': False,
            'error': None,
            'enter_err': '',
            'modify_err': ''
        }

        # 查询（通过退出码或 stdout 判断未找到）
        cmd_lookup = ["python", "info_lookup.py", "--gaijin", str(gaijin_id)]
        res_look = subprocess.run(cmd_lookup, capture_output=True, text=True, encoding="gbk")

        not_found = (res_look.returncode != 0) or ("未找到记录" in (res_look.stdout or ""))
        if not_found:
            cmd_enter = ["python", "info_enter.py", str(gaijin_id), str(name)]
            res_enter = subprocess.run(cmd_enter, capture_output=True, text=True, encoding="gbk")
            if res_enter.returncode != 0:
                row_status['error'] = '新增失败'
                row_status['enter_err'] = (res_enter.stderr or '').strip()
                logging.error(f"行={row_idx} id={gaijin_id} 操作=新增 结果=失败 错误={row_status['enter_err']}")
                continue
            row_status['created'] = True

            # 新增后再次确认记录已存在
            res_confirm = subprocess.run(cmd_lookup, capture_output=True, text=True, encoding="gbk")
            if res_confirm.returncode != 0 or ("未找到记录" in (res_confirm.stdout or "")):
                row_status['error'] = '新增未确认'
                logging.error(f"行={row_idx} id={gaijin_id} 操作=新增 结果=失败 原因=未确认")
                continue

        # 执行modify更新字段
        cmd_mod = ["python", "info_modify.py", str(gaijin_id)]

        def _has_value(v):
            return v is not None and str(v).strip() != ""

        if _has_value(state):
            cmd_mod += ["--state", str(state)]
        if _has_value(join_date):
            cmd_mod += ["--time", str(join_date)]
        if _has_value(landforce):
            cmd_mod += ["--landforce", str(landforce)]
        if _has_value(airforce):
            cmd_mod += ["--airforce", str(airforce)]
        if _has_value(navy):
            cmd_mod += ["--navy", str(navy)]

        # 如果没有任何要修改的字段，则跳过 modify
        if len(cmd_mod) > 1:
            res_mod = subprocess.run(cmd_mod, capture_output=True, text=True, encoding="gbk")
            if res_mod.returncode == 0:
                row_status['modified'] = True
            else:
                row_status['error'] = 'modify_failed'
                row_status['modify_err'] = (res_mod.stderr or '').strip()

        # 最终只输出一行结构化日志，描述本行执行结果（中文）
        action_parts = []
        if row_status['created']:
            action_parts.append('已创建')
        if row_status['modified']:
            action_parts.append('已修改')
        if not action_parts:
            action_parts.append('无变化')

        if row_status['error']:
            logging.error(f"行={row_idx} id={gaijin_id} 操作={'、'.join(action_parts)} 结果=失败 错误={row_status['error']} 新增错误={row_status['enter_err']} 修改错误={row_status['modify_err']}")
        else:
            logging.info(f"行={row_idx} id={gaijin_id} 操作={'、'.join(action_parts)} 结果=成功")

    logging.info(f"已经使用 {xlsx_path} 批量导入数据到数据库。")


def main():
    parser = argparse.ArgumentParser(description="从xlsx批量导入数据到sqlite")
    parser.add_argument('--xlsx', help='xlsx文件路径，默认为脚本同级目录下的data文件夹', default=None)
    args = parser.parse_args()

    # 配置日志记录
    logging.basicConfig(
        level=logging.INFO,
        handlers=[logging.FileHandler(log_path, encoding='utf-8')],
        format='%(asctime)s - [%(levelname)s] %(message)s',
        datefmt='%Y-%m-%dT%H:%M:%S'
    )
    if args.xlsx == None:
        logging.info("未指定xlsx文件路径，使用默认路径")
        args.xlsx = os.path.join(os.path.dirname(__file__), "MemberList.db 输入模板.xlsx")

    # 默认xlsx路径：脚本同级目录
    
    excel_to_db(args.xlsx, None, None)


if __name__ == "__main__":
    main()